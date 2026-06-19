import streamlit as st
import pandas as pd
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as xlImage
from openpyxl.utils import get_column_letter
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials
import smtplib
from email.message import EmailMessage

# ==========================================
# ⚙️ 정식 앱 이름 및 파이프 아이콘 고정 설정 (최상단 필수)
# ==========================================
st.set_page_config(
    page_title="GRE체크시트",
    page_icon="https://cdn-icons-png.flaticon.com/512/3252/3252917.png",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ==========================================
# ⚙️ 구글 클라우드 고유 ID 설정
# ==========================================
SPREADSHEET_ID ="1goTmdvN69Axic01bYIarMa3ej8V-pwEivZ9RLrIMTB4"

# 세션 상태 초기화
if 'align_data' not in st.session_state: st.session_state['align_data'] = []
if 'torque_data' not in st.session_state: st.session_state['torque_data'] = []
if 'photo1' not in st.session_state: st.session_state['photo1'] = None
if 'photo2' not in st.session_state: st.session_state['photo2'] = None
if 'photo3' not in st.session_state: st.session_state['photo3'] = None
if 'photo4' not in st.session_state: st.session_state['photo4'] = None

st.title("🛠️ GRE PIPE 모바일 체크시트")
st.markdown("---")

# ==========================================
# 📝 1. 보고서 현장 정보
# ==========================================
with st.expander("📄 보고서 현장 정보 (보고서 및 클라우드 동기화)", expanded=True):
    col_m1, col_m2 = st.columns(2)
    with col_m1:
        doc_date = st.date_input("작성일자", datetime.now())
        hull_no = st.text_input("호선 (Hull No.)", placeholder="예: SN2717")
        block_no = st.text_input("블록 (Block)", placeholder="예: D620S")
        tag_no = st.text_input("TAG NO.", placeholder="예: CG001")
        
    with col_m2:
        dept_name = st.selectbox("소속", ["선택 안함", "선행의장 1과", "1직1반", "1직2반", "2직1반", "3직1반", "3직2반", "4직2반", "4직3반", "대영기업"])
        doc_area = st.text_input("AREA", placeholder="예: 3001")
        doc_author = st.text_input("이름 (작성자)", value="")

    parts = ["SHI-GRE"]
    if hull_no: parts.append(hull_no.strip())
    if block_no: parts.append(block_no.strip())
    if tag_no: parts.append(tag_no.strip())
    
    if len(parts) == 1: doc_no = f"SHI-GRE-{datetime.now().strftime('%Y%m')}-001"
    else: doc_no = "-".join(parts)
        
    st.info(f"**📑 문서번호 (자동생성):** {doc_no}")

final_dept_name = "" if dept_name == "선택 안함" else dept_name

st.markdown("---")

# ==========================================
# 2. 기본 점검 사항
# ==========================================
st.header("1. 기본 점검 사항")
questions = [
    "파이프 작업 전 도면은 확인 하였는가?",
    "파이프 조인트부 선수, 선미에 각각 1개씩 SUPPORT는 설치되어있는가?",
    "파이프 내부에 이물질은 없는가?",
    "파이프 외부에 이물질이나 데미지는 없는가?",
    "파이프 설치 시 alignment 는 허용값안에 들어 오는가?",
    "파이프 설치 시 기준에 맞는 볼트 / 너트 / 와셔를 사용 하였는가?",
    "Earth Wire는 정확한 위치에 설치 하였는가?",
    "토크작업 전 토크렌치 교정검정일은 확인 하였는가? (교정성적서)",
    "볼트 체결시 메이커 매뉴얼 순서에 맞게 작업 하였는가?",
    "배관 커버링은 제대로 설치되었는가? (함석 + 난연성커버)"
]

basic_results = []
for i, q in enumerate(questions, 1):
    st.markdown(f"**Q{i}. {q}**")
    col1, col2 = st.columns([1, 1])
    with col1:
        status = st.radio("점검결과", ["양호", "불량", "해당없음"], key=f"status_{i}", horizontal=True, label_visibility="collapsed")
    with col2:
        remark = st.text_input("조치내용", placeholder="조치내용 입력...", key=f"remark_{i}", label_visibility="collapsed")
    basic_results.append({"순번": i, "항목": q, "결과": status, "비고": remark if remark else "-"})

overall_status = "양호"
for item in basic_results:
    if item["결과"] != "양호":
        overall_status = "불량"
        break

st.markdown("---")

# ==========================================
# 3. 데이터 입력 (얼라인먼트 & 토크) + 기준정보 팝업
# ==========================================
st.header("2. 데이터 입력")

# 💡 기준정보 팝업 기능 추가
with st.expander("📊 GRE PIPE 검사 기준표 보기 (클릭하여 펼치기)"):
    try:
        # 깃허브에 reference.jpg 파일이 있으면 불러옵니다.
        st.image("reference.jpg", caption="GRE PIPE 기준 데이터", use_container_width=True)
    except:
        st.info("💡 깃허브에 'reference.jpg'라는 이름으로 기준정보 이미지 파일을 올려두시면 이곳에 표시됩니다!")

st.info("💡 값을 입력한 후 [➕ 추가] 버튼을 안 눌러도, 입력된 값은 보고서에 자동으로 1줄 반영됩니다!")

dia_a = top = port = gap = coup_no = bottom = stb = ""
rem_a = "정상 범위 내"
dia_t = elem2 = elem1 = t_val = ""
serial = "P2024500178"

tab1, tab2 = st.tabs(["ALIGNMENT CHECK", "TORQUE CHECK"])

with tab1:
    col_a1, col_a2 = st.columns(2)
    with col_a1:
        dia_a = st.text_input("DIA (관경)", key="dia_a")
        top = st.text_input("TOP (상부)", key="top")
        port = st.text_input("PORT (좌현)", key="port")
        gap = st.text_input("GAP (간극)", key="gap")
    with col_a2:
        coup_no = st.text_input("COUPLING NUMBER", key="coup_no")
        bottom = st.text_input("BOTTOM (하부)", key="bottom")
        stb = st.text_input("STB (우현)", key="stb")
        rem_a = st.text_input("REMARK (비고)", key="rem_a", value="정상 범위 내")
    if st.button("➕ ALIGNMENT 리스트에 추가", use_container_width=True):
        st.session_state['align_data'].append([dia_a, coup_no, top, bottom, port, stb, gap, rem_a])
        st.success("추가되었습니다.")

with tab2:
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        dia_t = st.text_input("DIA (관경)", key="dia_t")
        elem2 = st.text_input("ELEM.NO (2)", key="elem2")
        serial = st.text_input("토크렌치 S/N", key="serial", value="P2024500178")
    with col_t2:
        elem1 = st.text_input("ELEM.NO (1)", key="elem1")
        t_val = st.text_input("TORQUE VALUE", key="t_val")
    if st.button("➕ TORQUE 리스트에 추가", use_container_width=True):
        st.session_state['torque_data'].append([dia_t, elem1, elem2, t_val, serial])
        st.success("추가되었습니다.")

st.markdown("---")

# ==========================================
# 4. 현장 증빙 사진 
# ==========================================
st.header("3. 토크렌치")
col_p1, col_p2 = st.columns(2)

with col_p1:
    st.markdown("**📸 토크렌치 교정번호**")
    up1 = st.file_uploader("업로드 1", type=["png", "jpg", "jpeg"], key="up1", label_visibility="collapsed")
    if up1:
        st.session_state['photo1'] = up1.getvalue()
        st.image(st.session_state['photo1'], use_container_width=True)
    else:
        st.session_state['photo1'] = None

with col_p2:
    st.markdown("**📸 토크렌치 세팅 값**")
    up2 = st.file_uploader("업로드 2", type=["png", "jpg", "jpeg"], key="up2", label_visibility="collapsed")
    if up2:
        st.session_state['photo2'] = up2.getvalue()
        st.image(st.session_state['photo2'], use_container_width=True)
    else:
        st.session_state['photo2'] = None

st.markdown("---")

st.header("4. 현장 사진")
col_p3, col_p4 = st.columns(2)

with col_p3:
    st.markdown("**📸 현장 사진 1**")
    up3 = st.file_uploader("업로드 3", type=["png", "jpg", "jpeg"], key="up3", label_visibility="collapsed")
    if up3:
        st.session_state['photo3'] = up3.getvalue()
        st.image(st.session_state['photo3'], use_container_width=True)
    else:
        st.session_state['photo3'] = None

with col_p4:
    st.markdown("**📸 현장 사진 2**")
    up4 = st.file_uploader("업로드 4", type=["png", "jpg", "jpeg"], key="up4", label_visibility="collapsed")
    if up4:
        st.session_state['photo4'] = up4.getvalue()
        st.image(st.session_state['photo4'], use_container_width=True)
    else:
        st.session_state['photo4'] = None

st.markdown("---")

final_align_data = st.session_state['align_data'].copy()
if not final_align_data and (dia_a or coup_no): 
    final_align_data = [[dia_a, coup_no, top, bottom, port, stb, gap, rem_a]]

final_torque_data = st.session_state['torque_data'].copy()
if not final_torque_data and (dia_t or elem1 or elem2):
    final_torque_data = [[dia_t, elem1, elem2, t_val, serial]]

# ==========================================
# 5. 엑셀 서식 생성 함수
# ==========================================
def generate_report(align_list, torque_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "품질검사보고서"
    ws.views.sheetView[0].showGridLines = False

    NAVY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    SUB_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    WHITE_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    SUB_FONT = Font(name="맑은 고딕", size=11, bold=True, color="1F4E78")
    BODY_FONT = Font(name="맑은 고딕", size=10)
    BODY_BOLD = Font(name="맑은 고딕", size=10, bold=True)
    ALERT_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FF0000")
    
    ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN_BORDER = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                         top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

    def apply_style(cell, font, align, border, fill=None):
        cell.font = font
        cell.alignment = align
        cell.border = border
        if fill: cell.fill = fill

    ws.merge_cells("A1:H2")
    title = ws["A1"]
    title.value = "GRE PIPE INSTALLATION & TORQUE INSPECTION REPORT"
    apply_style(title, Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF"), ALIGN_C, THIN_BORDER, NAVY_FILL)

    row_idx = 4
    meta_info = [
        ("작성일자", str(doc_date), "문서번호", doc_no),
        ("호선 (Hull No.)", hull_no, "블록 (Block)", block_no),
        ("TAG NO.", tag_no, "소속", final_dept_name),
        ("이름", doc_author, "AREA", doc_area)
    ]
    
    for lbl1, val1, lbl2, val2 in meta_info:
        ws.row_dimensions[row_idx].height = 22
        ws.cell(row=row_idx, column=1, value=lbl1); apply_style(ws.cell(row=row_idx, column=1), BODY_BOLD, ALIGN_C, THIN_BORDER, SUB_FILL)
        ws.cell(row=row_idx, column=5, value=lbl2); apply_style(ws.cell(row=row_idx, column=5), BODY_BOLD, ALIGN_C, THIN_BORDER, SUB_FILL)
        ws.merge_cells(f"B{row_idx}:D{row_idx}"); ws.cell(row=row_idx, column=2, value=val1)
        ws.merge_cells(f"F{row_idx}:H{row_idx}"); ws.cell(row=row_idx, column=6, value=val2)
        for c in range(2, 5): apply_style(ws.cell(row=row_idx, column=c), BODY_FONT, ALIGN_L, THIN_BORDER)
        for c in range(6, 9): apply_style(ws.cell(row=row_idx, column=c), BODY_FONT, ALIGN_L, THIN_BORDER)
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="1. 기본 점검 사항 (General Inspection)").font = SUB_FONT
    row_idx += 1
    
    headers_s1 = ["순번", "점검 항목 (Inspection Items)", "", "", "", "", "점검결과", "조치 내용 / 비고"]
    for col, h in enumerate(headers_s1, 1):
        apply_style(ws.cell(row=row_idx, column=col, value=h), WHITE_FONT, ALIGN_C, THIN_BORDER, NAVY_FILL)
    ws.merge_cells(f"B{row_idx}:F{row_idx}")
    
    for item in basic_results:
        row_idx += 1
        ws.row_dimensions[row_idx].height = 22
        ws.cell(row=row_idx, column=1, value=item["순번"])
        ws.cell(row=row_idx, column=2, value=item["항목"])
        ws.cell(row=row_idx, column=7, value=item["결과"])
        ws.cell(row=row_idx, column=8, value=item["비고"])
        for col in range(1, 9): apply_style(ws.cell(row=row_idx, column=col), BODY_FONT, ALIGN_C if col!=2 else ALIGN_L, THIN_BORDER)
        ws.merge_cells(f"B{row_idx}:F{row_idx}")

    row_idx += 1
    ws.merge_cells(f"A{row_idx}:H{row_idx}")
    alert = ws.cell(row=row_idx, column=1, value="※ 단 LEAK 발생시 담당PRO와 협의 후 MAXIMUM TORQUE값 사용가능")
    apply_style(alert, ALERT_FONT, ALIGN_C, THIN_BORDER)
        
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="2. ALIGNMENT CHECK RESULT").font = SUB_FONT
    row_idx += 1
    
    align_h = ["DIA (관경)", "COUPLING NO", "TOP (상부)", "BOTTOM (하부)", "PORT (좌현)", "STB (우현)", "GAP (간극)", "REMARK (비고)"]
    for col, h in enumerate(align_h, 1): 
        apply_style(ws.cell(row=row_idx, column=col, value=h), WHITE_FONT, ALIGN_C, THIN_BORDER, NAVY_FILL)
    for row_data in align_list:
        row_idx += 1
        for col, val in enumerate(row_data, 1): 
            apply_style(ws.cell(row=row_idx, column=col, value=val), BODY_FONT, ALIGN_C, THIN_BORDER)

    row_idx += 2
    ws.cell(row=row_idx, column=1, value="3. FLANGE JOINT TORQUE VALUE CHECK RESULT").font = SUB_FONT
    row_idx += 1
    
    torque_h = ["DIA (관경)", "ELEM.NO 1", "ELEM.NO 2", "토크 값 (N·m)", "토크렌치 S/N", "", "", ""]
    for col, h in enumerate(torque_h, 1):
        c = ws.cell(row=row_idx, column=col)
        apply_style(c, WHITE_FONT, ALIGN_C, THIN_BORDER, NAVY_FILL)
        if h: c.value = h
    ws.merge_cells(f"E{row_idx}:H{row_idx}")
    
    for row_data in torque_list:
        row_idx += 1
        for col, val in enumerate(row_data, 1): 
            apply_style(ws.cell(row=row_idx, column=col, value=val), BODY_FONT, ALIGN_C, THIN_BORDER)
        for col in range(6, 9):
            apply_style(ws.cell(row=row_idx, column=col), BODY_FONT, ALIGN_C, THIN_BORDER)
        ws.merge_cells(f"E{row_idx}:H{row_idx}")

    row_idx += 3
    ws.cell(row=row_idx, column=1, value="4. 현장 사진").font = SUB_FONT
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="[토크렌치 교정번호]"); apply_style(ws.cell(row=row_idx, column=1), BODY_BOLD, ALIGN_C, THIN_BORDER, SUB_FILL)
    ws.cell(row=row_idx, column=5, value="[토크렌치 세팅 값]"); apply_style(ws.cell(row=row_idx, column=5), BODY_BOLD, ALIGN_C, THIN_BORDER, SUB_FILL)
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws.merge_cells(f"E{row_idx}:H{row_idx}")
    for c in range(2, 5): ws.cell(row=row_idx, column=c).border = THIN_BORDER
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = THIN_BORDER
    
    row_idx += 1
    ws.row_dimensions[row_idx].height = 190
    for c in range(1, 9): ws.cell(row=row_idx, column=c).border = THIN_BORDER
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws.merge_cells(f"E{row_idx}:H{row_idx}")
    
    if st.session_state['photo1']:
        img1 = xlImage(io.BytesIO(st.session_state['photo1']))
        img1.width, img1.height = 310, 240
        ws.add_image(img1, f"A{row_idx}")
        
    if st.session_state['photo2']:
        img2 = xlImage(io.BytesIO(st.session_state['photo2']))
        img2.width, img2.height = 310, 240
        ws.add_image(img2, f"E{row_idx}")

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="[현장 사진 1]"); apply_style(ws.cell(row=row_idx, column=1), BODY_BOLD, ALIGN_C, THIN_BORDER, SUB_FILL)
    ws.cell(row=row_idx, column=5, value="[현장 사진 2]"); apply_style(ws.cell(row=row_idx, column=5), BODY_BOLD, ALIGN_C, THIN_BORDER, SUB_FILL)
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws.merge_cells(f"E{row_idx}:H{row_idx}")
    for c in range(2, 5): ws.cell(row=row_idx, column=c).border = THIN_BORDER
    for c in range(6, 9): ws.cell(row=row_idx, column=c).border = THIN_BORDER
    
    row_idx += 1
    ws.row_dimensions[row_idx].height = 190
    for c in range(1, 9): ws.cell(row=row_idx, column=c).border = THIN_BORDER
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    ws.merge_cells(f"E{row_idx}:H{row_idx}")
    
    if st.session_state['photo3']:
        img3 = xlImage(io.BytesIO(st.session_state['photo3']))
        img3.width, img3.height = 310, 240
        ws.add_image(img3, f"A{row_idx}")
        
    if st.session_state['photo4']:
        img4 = xlImage(io.BytesIO(st.session_state['photo4']))
        img4.width, img4.height = 310, 240
        ws.add_image(img4, f"E{row_idx}")

    widths = [13, 17, 13, 13, 13, 13, 13, 20]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# 7. 버튼 3개 분할 및 메일 다중 전송 시스템
# ==========================================
col_btn1, col_btn2, col_btn3 = st.columns(3)

with col_btn1:
    # (디자인 맞춤을 위해 빈 줄 하나 추가)
    st.markdown("<br>", unsafe_allow_html=True) 
    if st.button("🚀 데이터 전송", type="primary", use_container_width=True):
        with st.spinner("구글 시트에 스마트 요약 누적 중..."):
            try:
                gcp_creds = json.loads(st.secrets["gcp_service_account"])
                if "private_key" in gcp_creds:
                    gcp_creds["private_key"] = gcp_creds["private_key"].replace("\\n", "\n")
                creds = Credentials.from_service_account_info(gcp_creds, scopes=["https://www.googleapis.com/auth/spreadsheets"])
                gc = gspread.authorize(creds)
                sh = gc.open_by_key(SPREADSHEET_ID)
                
                formatted_date = doc_date.strftime('%y-%m-%d')
                photo_status = "유" if (st.session_state['photo1'] or st.session_state['photo2'] or st.session_state['photo3'] or st.session_state['photo4']) else "무"
                summary_row = [formatted_date, hull_no, block_no, tag_no, final_dept_name, doc_author, overall_status, photo_status]
                
                sh.worksheet("기본점검로그").append_row(summary_row)
                if final_align_data: sh.worksheet("얼라인먼트로그").append_rows([[formatted_date, hull_no, tag_no] + row for row in final_align_data])
                if final_torque_data: sh.worksheet("토크로그").append_rows([[formatted_date, hull_no, tag_no] + row for row in final_torque_data])
                
                st.success("🎉 구글 시트 데이터 요약 누적 완료!")
            except Exception as e:
                st.error(f"시트 전송 실패: {e}")

with col_btn2:
    st.markdown("<br>", unsafe_allow_html=True) 
    try:
        excel_bytes = generate_report(final_align_data, final_torque_data)
        file_name = f"{doc_no}.xlsx"
        st.download_button(
            label="📥 양식 다운로드(Excel)",
            data=excel_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    except Exception as e:
        st.warning("보고서 엑셀을 생성하는 중 오류가 발생했습니다.")

with col_btn3:
    # 💡 담당자 이메일 목록 딕셔너리
    email_dict = {
        "홍의현 프로 (본인)": "h9222.hong@samsung.com",
        "김민효 프로": "minhyo7.kim@samsung.com",
        "최지원 프로": "jw5241.choi@samsung.com",
        "황보현 프로": "bh0623.hwang@samsung.com",
        "전민재 프로": "minjae3.jeon@samsung.com"
    }
    
    # 💡 다중 선택창 (버튼 바로 위에 배치)
    selected_receivers = st.multiselect(
        "수신자 선택",
        options=list(email_dict.keys()),
        default=["홍의현 프로 (본인)"],
        label_visibility="collapsed"
    )
    
    if st.button("📧 메일 전송", use_container_width=True):
        if not selected_receivers:
            st.warning("메일을 받을 사람을 1명 이상 선택해 주세요!")
        else:
            with st.spinner("사내 메일로 엑셀 보고서를 전송 중..."):
                try:
                    excel_bytes = generate_report(final_align_data, final_torque_data)
                    file_name = f"{doc_no}.xlsx"
                    
                    sender_email = st.secrets["email"]["sender_email"]
                    app_pw = st.secrets["email"]["app_password"]
                    
                    # 선택된 사람들의 실제 이메일 주소만 추출
                    target_emails = [email_dict[name] for name in selected_receivers]
                    
                    msg = EmailMessage()
                    msg['Subject'] = f'[품질검사보고서] {doc_no} 검사 완료 건'
                    msg['From'] = sender_email
                    msg['To'] = ", ".join(target_emails) # 💡 여러 명에게 동시 전송
                    msg.set_content(f"현장에서 작성된 GRE PIPE 품질검사보고서 ({file_name}) 첨부드립니다.\n\n- 점검구역: {doc_area}\n- 점검결과: {overall_status}\n\n*본 메일은 모바일 체크시트 시스템에서 자동 발송되었습니다.")
                    
                    msg.add_attachment(
                        excel_bytes, 
                        maintype='application', 
                        subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                        filename=file_name
                    )
                    
                    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                        smtp.login(sender_email, app_pw)
                        smtp.send_message(msg)
                        
                    st.success(f"🎉 선택하신 {len(target_emails)}명에게 메일 전송 완료!")
                    
                except KeyError:
                    st.error("메일 발송 실패: Streamlit Secrets에 [email] 설정을 확인해 주세요!")
                except Exception as e:
                    st.error(f"메일 발송 중 오류가 발생했습니다: {e}")

st.markdown("---")
if st.button("🔄 새로운 점검 시작 (입력창 초기화)", use_container_width=True):
    st.session_state['align_data'] = []
    st.session_state['torque_data'] = []
    st.session_state['photo1'] = None
    st.session_state['photo2'] = None
    st.session_state['photo3'] = None
    st.session_state['photo4'] = None
    st.rerun()
